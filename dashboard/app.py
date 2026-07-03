import os
import sys
import hmac
import datetime
import csv
import json
import io
import urllib.request
import urllib.parse
from flask import Flask, jsonify, request, send_file, render_template, Response

# Ensure workspace root is in python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from database.sqlite_db import SQLiteDatabase
from reports.reporter import AttendanceReporter, format_seconds, time_to_seconds, calculate_time_diff_seconds
from bot.shifts import get_employee_shift

app = Flask(__name__,
            static_folder="static",
            template_folder="templates")

# This dashboard exposes employee PII, fines, and admin actions (ban, edit
# attendance, approve/reject requests) with no other access control. Set
# DASHBOARD_USERNAME / DASHBOARD_PASSWORD when hosting this publicly (e.g. on
# Render) to gate it behind HTTP Basic Auth. Left unset, auth stays off so
# local development is unaffected.
DASHBOARD_USERNAME = os.getenv("DASHBOARD_USERNAME", "admin")
DASHBOARD_PASSWORD = os.getenv("DASHBOARD_PASSWORD", "")


@app.before_request
def _require_dashboard_auth():
    if not DASHBOARD_PASSWORD:
        return None
    auth = request.authorization
    valid = bool(auth) and hmac.compare_digest(auth.username or "", DASHBOARD_USERNAME) \
        and hmac.compare_digest(auth.password or "", DASHBOARD_PASSWORD)
    if not valid:
        return Response(
            "Authentication required.", 401,
            {"WWW-Authenticate": 'Basic realm="Attendance Dashboard"'}
        )
    return None

# Initialize database connection (self-bootstrapping: creates the schema if this
# is the first process to touch a fresh DB_PATH, so the dashboard doesn't depend
# on the bot having run first).
db = SQLiteDatabase(config.DB_PATH)
db.connect()
_schema_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "schema.sql")
db.execute_schema(_schema_path)

# Setup system timezone
tz = config.TIMEZONE

# Safe fallback shift defaults
DEFAULT_SHIFT_START = "09:00:00"
DEFAULT_SHIFT_END = "18:00:00"

def notify_employee(telegram_id: int, message: str, parse_mode: str = "HTML") -> bool:
    """Send a private Telegram message directly to an employee.
    
    NOTE: This only works if the employee has previously sent a message to the bot
    (i.e. pressed /start). Telegram blocks bots from initiating conversations.
    Returns True on success, False on any failure.
    """
    url = f"https://api.telegram.org/bot{config.TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": telegram_id,
        "text": message,
        "parse_mode": parse_mode
    }
    try:
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            if res_data.get('ok'):
                print(f"[NOTIFY] ✅ Message sent to telegram_id={telegram_id}")
                return True
            else:
                print(f"[NOTIFY] ❌ Telegram rejected for {telegram_id}: {res_data}")
                return False
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        print(f"[NOTIFY] ❌ HTTP {e.code} for telegram_id={telegram_id}: {body}")
        return False
    except Exception as e:
        print(f"[NOTIFY] ❌ Exception for telegram_id={telegram_id}: {e}")
        return False

def send_and_record_permission_notification(req: dict, status: str, approver_name: str) -> bool:
    """Helper to compile, send and record a permission request notification."""
    try:
        from bot.permission_handler import REQUEST_TYPE_LABELS
        from reports.reporter import format_seconds
        import datetime as _notification_dt
        label = REQUEST_TYPE_LABELS.get(req['request_type'], req['request_type'])
        
        # Build decided_at display
        decided_at = req.get('decided_at') or _notification_dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        decided_display = decided_at[:16] if decided_at else '—'
        
        if status == 'approved':
            message = (
                f"✅ *Permission Request Approved*\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"📌 Type: {label}\n"
                f"📅 Request Date: `{req['date']}`\n"
                f"⏰ Time Range: `{req['start_time']}` → `{req['end_time']}`\n"
                f"⏱️ Duration: `{format_seconds(req['duration_seconds'])}`\n"
                f"📝 Reason: {req['reason']}\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"✅ *Status: APPROVED*\n"
                f"👤 Approved by: *{approver_name}*\n"
                f"🕐 Decided at: `{decided_display}`\n\n"
                f"_Your approved hours have been credited to your attendance record._"
            )
        else:  # rejected
            message = (
                f"❌ *Permission Request Rejected*\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"📌 Type: {label}\n"
                f"📅 Request Date: `{req['date']}`\n"
                f"⏰ Time Range: `{req['start_time']}` → `{req['end_time']}`\n"
                f"📝 Reason: {req['reason']}\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"❌ *Status: REJECTED*\n"
                f"👤 Reviewed by: *{approver_name}*\n"
                f"🕐 Decided at: `{decided_display}`\n\n"
                f"_If you have questions, please contact your supervisor._"
            )
        
        success = notify_employee(req['telegram_id'], message, parse_mode="Markdown")
        db.update_permission_notification_status(req['id'], 'sent' if success else 'failed')
        return success
    except Exception as err:
        print(f"⚠️ Error preparing notification for request {req.get('id')}: {err}")
        try:
            db.update_permission_notification_status(req['id'], 'failed')
        except Exception:
            pass
        return False

def resolve_shift_name(start_time: str, end_time: str) -> str:
    """Helper to classify shift timing details into readable descriptions."""
    if start_time == "08:30:00" and end_time == "19:00:00":
        return "Day Shift A (08:30 - 19:00)"
    elif start_time == "09:30:00" and end_time == "20:00:00":
        return "Day Shift B (09:30 - 20:00)"
    elif start_time == "08:30:00" and end_time == "20:30:00":
        return "Extended Day Shift (08:30 - 20:30)"
    elif start_time == "09:00:00" and end_time == "19:30:00":
        return "Mid Shift (09:00 - 19:30)"
    elif start_time == "10:00:00" and end_time == "20:30:00":
        return "Late Shift (10:00 - 20:30)"
    elif start_time == "09:00:00" and end_time == "18:00:00":
        return "Standard Office Shift (09:00 - 18:00)"
    else:
        return f"Custom Shift ({start_time} - {end_time})"

@app.route('/')
def index():
    """Serves the main single page dashboard."""
    return render_template('index.html')

@app.route('/api/summary', methods=['GET'])
def get_summary():
    """API for overview card counters."""
    today_str = datetime.datetime.now(tz).strftime("%Y-%m-%d")
    users = db.get_all_users()
    
    total_employees = len([u for u in users if u['role'] == 'employee' or (u['username'] and u['username'].lower() == 'professor_noxx')])
    banned_employees = len([u for u in users if u['status'] == 'banned' and (u['role'] == 'employee' or (u['username'] and u['username'].lower() == 'professor_noxx'))])
    
    active_now = 0
    on_break = 0
    
    for u in users:
        # Check active logins
        sess = db.get_active_attendance_session(u['telegram_id'])
        if sess:
            active_now += 1
            # Check if active session is on break
            brk = db.get_active_break_session(u['telegram_id'])
            if brk:
                on_break += 1
                
    # Fines stats
    conn = db.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT SUM(amount) FROM fines WHERE date = ?", (today_str,))
    fines_today = cursor.fetchone()[0] or 0.0
    
    # Pending requests count (early logout)
    cursor.execute("SELECT COUNT(*) FROM early_logout_requests WHERE status = 'pending'")
    pending_requests = cursor.fetchone()[0] or 0

    # Pending permission requests count
    pending_permissions = 0
    try:
        cursor.execute("SELECT COUNT(*) FROM permission_requests WHERE status = 'pending'")
        pending_permissions = cursor.fetchone()[0] or 0
    except Exception:
        pass
    
    return jsonify({
        "total_employees": total_employees,
        "active_employees": active_now,
        "on_break": on_break,
        "banned_employees": banned_employees,
        "fines_today": fines_today,
        "pending_requests": pending_requests,
        "pending_permissions": pending_permissions,
    })

@app.route('/api/employees', methods=['GET'])
def get_employees():
    """Returns the lists of all employees with their current active statuses and details."""
    from bot.shifts import EMPLOYEE_DATA
    today_str = datetime.datetime.now(tz).strftime("%Y-%m-%d")
    users = db.get_all_users()
    
    results = []
    bad_names = {
        "login.", "login", "log in", "logout.", "logout", "log out", 
        "break out.", "break out", "breakin.", "break in.", "break in", 
        "break in ☕", "out.", "out", "in.", "in", "login 🟢",
        "lunch out.", "lunch out", "lunchin.", "lunch in.", "lunch in"
    }
    
    for u in users:
        if u['role'] == 'admin' and (not u['username'] or u['username'].lower() != 'professor_noxx'):
            continue
            
        telegram_id = u['telegram_id']
        username = u['username']
        uname_lower = username.lower() if username else ""
        meta = EMPLOYEE_DATA.get(uname_lower) if uname_lower else None
        
        summary = AttendanceReporter.get_employee_daily_summary(db, telegram_id, today_str)
        
        # Determine current action status
        active_sess = db.get_active_attendance_session(telegram_id)
        active_break = db.get_active_break_session(telegram_id)
        active_move = db.get_active_in_out_session(telegram_id)
        
        # Check if fine is applied for today's session
        fine_applied = 0
        fine_amount = 0.0
        is_half_day = 0
        has_session_today = False
        
        if active_sess:
            has_session_today = True
            if active_break:
                current_status = "On Break"
            elif active_move:
                current_status = "Field Visit"
            else:
                current_status = "Working"
                
            # check the active or last session fine values
            cursor = db.connect().cursor()
            cursor.execute("SELECT fine_applied, fine_amount, is_half_day FROM attendance_sessions WHERE id = ?", (active_sess['id'],))
            row = cursor.fetchone()
            if row:
                fine_applied, fine_amount, is_half_day = row
        else:
            # check if there's any completed session today
            conn = db.connect()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT fine_applied, fine_amount, is_half_day FROM attendance_sessions WHERE telegram_id = ? AND date = ? ORDER BY id DESC LIMIT 1",
                (telegram_id, today_str)
            )
            row = cursor.fetchone()
            if row:
                fine_applied, fine_amount, is_half_day = row
                has_session_today = True
                current_status = "Offline"
            else:
                has_session_today = False
                current_status = "Absent"
                
        # Check early logout request
        req = db.get_early_logout_request_by_date(telegram_id, today_str)
        req_status = req['status'] if req else 'none'
        
        # Clean full_name if bad name
        full_name = u['full_name']
        if meta and (not full_name or full_name.strip().lower() in bad_names):
            full_name = meta.get("name") or full_name
            
        # Fallback fields from EMPLOYEE_DATA
        project_val = u.get("project")
        if not project_val and meta:
            project_val = meta.get("project")
        project_val = project_val or ""
        
        # Determine shift type
        shift_type_val = u.get("shift_type")
        if not shift_type_val and meta:
            shift_type_val = meta.get("shift_type")
        shift_type_val = shift_type_val or "day"
        
        s_start = u['shift_start']
        if not s_start and meta:
            s_start = meta.get("shift_start")
        s_start = s_start or DEFAULT_SHIFT_START
        
        s_end = u['shift_end']
        if not s_end and meta:
            s_end = meta.get("shift_end")
        s_end = s_end or DEFAULT_SHIFT_END
        
        results.append({
            "telegram_id": telegram_id,
            "username": username or "NoUsername",
            "full_name": full_name,
            "shift_name": resolve_shift_name(s_start, s_end),
            "shift_start": s_start,
            "shift_end": s_end,
            "status": u['status'] or 'active',  # 'active' or 'banned'
            "current_status": current_status,
            "has_session_today": has_session_today,
            "net_working_str": summary['net_working_str'],
            "total_login_str": summary['total_login_str'],
            "fine_applied": fine_applied,
            "fine_amount": fine_amount,
            "is_half_day": is_half_day,
            "request_status": req_status,
            "employee_id": u.get("employee_id") or "",
            "project": project_val,
            "shift_type": shift_type_val,
            "break_allowance": u.get("break_allowance") or 65,
            "attendance_settings": u.get("attendance_settings") or ""
        })
        
    return jsonify(results)

@app.route('/api/shifts/employees', methods=['GET'])
def get_shifts_employees():
    """
    Returns a merged list of all database users and employees defined in EMPLOYEE_DATA
    along with their current attendance status today, working hours, and registration info.
    """
    from bot.shifts import EMPLOYEE_DATA
    today_str = datetime.datetime.now(tz).strftime("%Y-%m-%d")
    
    # Get registered users from database
    users = db.get_all_users()
    users_by_username = {}
    for u in users:
        if u['role'] == 'admin' and (not u['username'] or u['username'].lower() != 'professor_noxx'):
            continue
        if u['username']:
            users_by_username[u['username'].lower()] = u
            
    # Merge phase: build a uniform mapping of all employees (both hardcoded and dynamic)
    all_employees_meta = {}
    
    # 1. Start with hardcoded employees in EMPLOYEE_DATA
    for username, meta in EMPLOYEE_DATA.items():
        uname_lower = username.lower()
        db_user = users_by_username.get(uname_lower)
        if db_user:
            all_employees_meta[uname_lower] = {
                "name": db_user["full_name"],
                "project": db_user.get("project") or meta.get("project") or "1",
                "shift_start": db_user["shift_start"] or meta.get("shift_start") or "09:00:00",
                "shift_end": db_user["shift_end"] or meta.get("shift_end") or "18:00:00",
                "shift_type": db_user.get("shift_type") or meta.get("shift_type") or "day",
                "db_user": db_user
            }
        else:
            all_employees_meta[uname_lower] = {
                "name": meta["name"],
                "project": meta["project"],
                "shift_start": meta["shift_start"],
                "shift_end": meta["shift_end"],
                "shift_type": meta["shift_type"],
                "db_user": None
            }
            
    # 2. Add dynamic/registered users from the database that are NOT in EMPLOYEE_DATA
    for u in users:
        if u['role'] == 'admin' and (not u['username'] or u['username'].lower() != 'professor_noxx'):
            continue
        uname_lower = u['username'].lower() if u['username'] else None
        
        is_processed = False
        if uname_lower and uname_lower in all_employees_meta:
            is_processed = True
            
        if not is_processed:
            key = uname_lower if uname_lower else f"id_{u['telegram_id']}"
            all_employees_meta[key] = {
                "name": u["full_name"],
                "project": u.get("project") or "Unassigned",
                "shift_start": u["shift_start"] or DEFAULT_SHIFT_START,
                "shift_end": u["shift_end"] or DEFAULT_SHIFT_END,
                "shift_type": u.get("shift_type") or "day",
                "db_user": u
            }
            
    results = []
    
    for identifier, meta in all_employees_meta.items():
        db_user = meta["db_user"]
        
        telegram_id = None
        role = "employee"
        status = "unregistered"
        current_status = "Absent"
        net_working_str = "00:00:00"
        total_login_str = "00:00:00"
        total_break_str = "00:00:00"
        total_move_str = "00:00:00"
        fine_applied = 0
        fine_amount = 0.0
        fine_reason = ""
        is_half_day = 0
        
        if db_user:
            telegram_id = db_user['telegram_id']
            role = db_user['role']
            status = db_user['status'] or 'active'  # 'active' or 'banned'
            
            # If the telegram_id is negative, it's a placeholder (unregistered on bot but pre-created by admin)
            if telegram_id is not None and telegram_id < 0:
                status = "unregistered"
                current_status = "Absent"
            elif telegram_id is not None and telegram_id > 0:
                try:
                    summary = AttendanceReporter.get_employee_daily_summary(db, telegram_id, today_str)
                    net_working_str = summary['net_working_str']
                    total_login_str = summary['total_login_str']
                    total_break_str = summary['total_break_str']
                    total_move_str = summary['total_move_str']
                    
                    active_sess = db.get_active_attendance_session(telegram_id)
                    active_break = db.get_active_break_session(telegram_id)
                    active_move = db.get_active_in_out_session(telegram_id)
                    
                    if status == "banned":
                        current_status = "Banned"
                    elif active_sess:
                        if active_break:
                            current_status = "On Break"
                        elif active_move:
                            current_status = "Field Visit"
                        else:
                            current_status = "Working"
                    else:
                        conn = db.connect()
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT is_half_day, fine_applied, fine_amount, fine_reason FROM attendance_sessions WHERE telegram_id = ? AND date = ? ORDER BY id DESC LIMIT 1",
                            (telegram_id, today_str)
                        )
                        row = cursor.fetchone()
                        if row:
                            is_half_day, fine_applied, fine_amount, fine_reason = row
                            current_status = "Half Day" if is_half_day else "Full Day"
                        else:
                            current_status = "Absent"
                except Exception as e:
                    print(f"⚠️ Error fetching data for telegram_id {telegram_id}: {e}")
                    current_status = "Absent"
        
        username_val = identifier if not identifier.startswith("id_") else ""
        if db_user and db_user['username']:
            username_val = db_user['username']
            
        # Safe defaults for shift times
        s_start = meta["shift_start"] or DEFAULT_SHIFT_START
        s_end = meta["shift_end"] or DEFAULT_SHIFT_END
        s_type = meta["shift_type"] or "day"
        s_project = meta["project"] or "Unassigned"
        
        results.append({
            "telegram_id": telegram_id,
            "username": username_val,
            "full_name": meta["name"],
            "project": s_project,
            "shift_name": f"{'Day' if s_type == 'day' else 'Night'} Shift ({s_start[:5]} - {s_end[:5]})",
            "shift_start": s_start,
            "shift_end": s_end,
            "shift_type": s_type,
            "status": status, # 'unregistered', 'active', or 'banned'
            "current_status": current_status, # 'Working', 'On Break', 'Field Visit', 'Absent', 'Half Day', 'Full Day', 'Banned'
            "net_working_str": net_working_str,
            "total_login_str": total_login_str,
            "total_break_str": total_break_str,
            "total_move_str": total_move_str,
            "fine_applied": fine_applied,
            "fine_amount": fine_amount,
            "fine_reason": fine_reason or "",
            "is_half_day": is_half_day
        })
        
    return jsonify(results)

@app.route('/api/employee/<telegram_id>', methods=['GET'])
def get_employee_detail(telegram_id):
    """Retrieve full employee profile history, growth, and metrics."""
    telegram_id = int(telegram_id)
    user = db.get_user(telegram_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
        
    conn = db.connect()
    cursor = conn.cursor()
    
    # 1. Fetch attendance logs
    cursor.execute("SELECT * FROM attendance_sessions WHERE telegram_id = ? ORDER BY date DESC, login_time DESC", (telegram_id,))
    att_rows = [dict(row) for row in cursor.fetchall()]
    
    # 2. Fetch break logs
    cursor.execute("SELECT * FROM break_sessions WHERE telegram_id = ? ORDER BY date DESC, break_in_time DESC", (telegram_id,))
    brk_rows = [dict(row) for row in cursor.fetchall()]
    
    # 3. Fetch movement logs
    cursor.execute("SELECT * FROM in_out_sessions WHERE telegram_id = ? ORDER BY date DESC, in_time DESC", (telegram_id,))
    move_rows = [dict(row) for row in cursor.fetchall()]
    
    # 4. Fetch fines
    cursor.execute("SELECT * FROM fines WHERE telegram_id = ? ORDER BY date DESC", (telegram_id,))
    fines_rows = [dict(row) for row in cursor.fetchall()]
    
    # Calculate performance stats (last 30 days)
    total_login_sec = 0
    total_break_sec = 0
    days_present = set()
    total_fines_amount = sum([f['amount'] for f in fines_rows])
    half_days_count = 0
    
    daily_durations = {}
    
    for row in att_rows:
        date = row['date']
        days_present.add(date)
        duration = row['duration'] or 0
        total_login_sec += duration
        if row['is_half_day']:
            half_days_count += 1
            
        daily_durations[date] = daily_durations.get(date, 0) + duration

    for row in brk_rows:
        date = row['date']
        duration = row['duration'] or 0
        total_break_sec += duration
        if date in daily_durations:
            daily_durations[date] = max(0, daily_durations[date] - duration)

    # Net hours
    net_working_sec = max(0, total_login_sec - total_break_sec)
    unique_days_count = len(days_present)
    avg_daily_seconds = net_working_sec // unique_days_count if unique_days_count > 0 else 0
    
    # Productivity rate: Average daily hours compared to full 10-hour standard (36000 seconds)
    productivity_percent = min(100, int((avg_daily_seconds / 36000.0) * 100)) if avg_daily_seconds > 0 else 0
    
    # growth stats: compare past 7 days average vs previous 7 days average
    today = datetime.date.today()
    last_7_days = [str(today - datetime.timedelta(days=i)) for i in range(7)]
    prev_7_days = [str(today - datetime.timedelta(days=i)) for i in range(7, 14)]
    
    avg_last_7 = sum([daily_durations.get(d, 0) for d in last_7_days]) / 7.0
    avg_prev_7 = sum([daily_durations.get(d, 0) for d in prev_7_days]) / 7.0
    
    growth_rate = 0.0
    if avg_prev_7 > 0:
        growth_rate = ((avg_last_7 - avg_prev_7) / avg_prev_7) * 100.0
        
    return jsonify({
        "profile": {
            "telegram_id": user['telegram_id'],
            "username": user['username'] or "NoUsername",
            "full_name": user['full_name'],
            "role": user['role'],
            "shift_name": resolve_shift_name(user['shift_start'], user['shift_end']),
            "shift_start": user['shift_start'],
            "shift_end": user['shift_end'],
            "status": user['status'],
            "registered_at": user['registered_at']
        },
        "stats": {
            "days_present": unique_days_count,
            "total_working_hours": format_seconds(net_working_sec),
            "avg_working_hours": format_seconds(avg_daily_seconds),
            "total_fines": total_fines_amount,
            "half_days": half_days_count,
            "productivity_score": productivity_percent,
            "growth_trend": round(growth_rate, 1)
        },
        "attendance_logs": att_rows,
        "break_logs": brk_rows,
        "movement_logs": move_rows,
        "fines_logs": fines_rows
    })

@app.route('/api/employees', methods=['POST'])
def add_employee():
    """Create a new employee record. Telegram ID is required as the primary identifier."""
    data = request.get_json() or {}
    full_name = data.get("full_name")
    username = data.get("username")
    employee_id = data.get("employee_id")   # Optional
    project = data.get("project")
    shift_type = data.get("shift_type", "day")
    shift_start = data.get("shift_start", "09:00:00")
    shift_end = data.get("shift_end", "18:00:00")
    break_allowance = data.get("break_allowance", 65)
    attendance_settings = data.get("attendance_settings", "")
    telegram_id_val = data.get("telegram_id")

    if not full_name:
        return jsonify({"error": "Full Name is required"}), 400

    # Telegram ID is the primary unique identifier — it is required
    if telegram_id_val is None or str(telegram_id_val).strip() == "":
        return jsonify({"error": "Telegram User ID is required. It must match the user's actual Telegram account ID."}), 400
    try:
        telegram_id = int(telegram_id_val)
    except (ValueError, TypeError):
        return jsonify({"error": "Telegram ID must be a numeric integer"}), 400

    # Clean username
    if username:
        username = username.strip().replace("@", "")

    # Employee ID is optional — only check uniqueness if one is provided
    if employee_id:
        conn = db.connect()
        cursor = conn.cursor()
        # Allow the same employee_id if it belongs to the same telegram_id (upsert case)
        cursor.execute("SELECT telegram_id FROM users WHERE employee_id = ? AND telegram_id != ?", (employee_id, telegram_id))
        if cursor.fetchone():
            return jsonify({"error": f"Employee ID '{employee_id}' is already assigned to another employee"}), 400

    # Check if Telegram ID already exists — if so, update the existing record (upsert)
    existing = db.get_user(telegram_id)
    if existing:
        # Update the existing record with the new details instead of blocking
        success = db.update_staff_user(
            telegram_id=telegram_id,
            username=username,
            full_name=full_name,
            employee_id=employee_id,
            project=project,
            shift_type=shift_type,
            shift_start=shift_start,
            shift_end=shift_end,
            break_allowance=int(break_allowance),
            attendance_settings=attendance_settings,
            status=data.get("status", existing.get("status", "active"))
        )
        if success:
            return jsonify({"success": True, "telegram_id": telegram_id, "updated": True})
        else:
            return jsonify({"error": "Failed to update existing employee record."}), 500

    # Create the new user record
    success = db.create_staff_user(
        telegram_id=telegram_id,
        username=username,
        full_name=full_name,
        employee_id=employee_id,
        project=project,
        shift_type=shift_type,
        shift_start=shift_start,
        shift_end=shift_end,
        break_allowance=int(break_allowance),
        attendance_settings=attendance_settings
    )

    if success:
        return jsonify({"success": True, "telegram_id": telegram_id})
    else:
        return jsonify({"error": "Failed to create employee record. Possibly a duplicate Telegram ID or username."}), 500


@app.route('/api/employees/<telegram_id>', methods=['PUT'])
def edit_employee(telegram_id):
    """Edit an existing employee record."""
    telegram_id = int(telegram_id)
    data = request.get_json() or {}
    
    # Handle optional Telegram ID modification
    new_telegram_id_val = data.get("telegram_id")
    new_telegram_id = None
    if new_telegram_id_val is not None and str(new_telegram_id_val).strip() != "":
        try:
            new_telegram_id = int(new_telegram_id_val)
        except ValueError:
            return jsonify({"error": "Telegram ID must be a numeric integer"}), 400

    # Verify if user exists under original ID
    user = db.get_user(telegram_id)
    if not user:
        return jsonify({"error": "Employee not found"}), 404

    # If updating Telegram ID, perform migration across tables
    if new_telegram_id and new_telegram_id != telegram_id:
        existing = db.get_user(new_telegram_id)
        if existing:
            return jsonify({"error": f"Telegram ID {new_telegram_id} is already in use by {existing['full_name']}."}), 400
        
        if not db.update_telegram_id(telegram_id, new_telegram_id):
            return jsonify({"error": "Failed to update Telegram ID. Possibly duplicate ID."}), 500
        telegram_id = new_telegram_id

    full_name = data.get("full_name")
    username = data.get("username")
    employee_id = data.get("employee_id")
    project = data.get("project")
    shift_type = data.get("shift_type", "day")
    shift_start = data.get("shift_start", "09:00:00")
    shift_end = data.get("shift_end", "18:00:00")
    break_allowance = data.get("break_allowance", 65)
    attendance_settings = data.get("attendance_settings", "")
    status = data.get("status", "active")

    if not full_name:
        return jsonify({"error": "Full Name is required"}), 400

    # Clean username
    if username:
        username = username.strip().replace("@", "")

    # If employee_id is changing, check if unique
    if employee_id:
        conn = db.connect()
        cursor = conn.cursor()
        cursor.execute("SELECT telegram_id FROM users WHERE employee_id = ? AND telegram_id != ?", (employee_id, telegram_id))
        if cursor.fetchone():
            return jsonify({"error": f"Employee ID '{employee_id}' is already assigned to another employee"}), 400

    success = db.update_staff_user(
        telegram_id=telegram_id,
        username=username,
        full_name=full_name,
        employee_id=employee_id,
        project=project,
        shift_type=shift_type,
        shift_start=shift_start,
        shift_end=shift_end,
        break_allowance=int(break_allowance),
        attendance_settings=attendance_settings,
        status=status
    )

    if success:
        return jsonify({"success": True})
    else:
        return jsonify({"error": "Failed to update employee details"}), 500


@app.route('/api/employees/<telegram_id>', methods=['DELETE'])
def delete_employee(telegram_id):
    """Delete an employee record."""
    telegram_id = int(telegram_id)
    user = db.get_user(telegram_id)
    if not user:
        return jsonify({"error": "Employee not found"}), 404

    success = db.delete_user(telegram_id)
    if success:
        return jsonify({"success": True})
    else:
        return jsonify({"error": "Failed to delete employee"}), 500


@app.route('/api/employees/<telegram_id>/status', methods=['POST'])
def update_status(telegram_id):
    """Ban or reactivate an employee."""
    telegram_id = int(telegram_id)
    data = request.get_json() or {}
    new_status = data.get("status")
    if new_status not in ["active", "banned"]:
        return jsonify({"error": "Invalid status"}), 400
        
    success = db.update_user_status(telegram_id, new_status)
    if success:
        if new_status == "banned":
            notify_employee(telegram_id, "🚫 *Access Denied*: You have been banned from the attendance system by the administrator.", parse_mode="Markdown")
        else:
            notify_employee(telegram_id, "✅ *Access Reinstated*: Your attendance tracking account has been activated.", parse_mode="Markdown")
        return jsonify({"success": True})
    return jsonify({"error": "Failed to update status"}), 500

@app.route('/api/attendance/<int:session_id>/fine', methods=['POST'])
def apply_session_fine(session_id):
    """Applies or toggles a fine associated with a specific attendance record session."""
    data = request.get_json() or {}
    fine_applied = 1 if data.get("fine_applied") else 0
    fine_amount = float(data.get("fine_amount", 500.0))
    fine_reason = data.get("fine_reason", "Policy Violation")
    remarks = data.get("remarks", "")
    
    # Fetch attendance session to get employee ID
    conn = db.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT telegram_id, date, name FROM attendance_sessions WHERE id = ?", (session_id,))
    sess = cursor.fetchone()
    if not sess:
        return jsonify({"error": "Attendance session not found"}), 404
        
    telegram_id, date, name = sess
    
    # Update session fine flags
    success = db.set_attendance_fine(session_id, fine_applied, fine_amount, fine_reason)
    if not success:
        return jsonify({"error": "Database write error"}), 500
        
    # Synchronize audit fines table
    if fine_applied:
        db.create_fine(telegram_id, date, fine_amount, fine_reason)
        # Format current timestamp for notification
        current_time_str = datetime.datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")
        import html
        escaped_reason = html.escape(fine_reason)
        escaped_remarks = html.escape(remarks) if remarks else ""
        remarks_block = f"\n💬 <b>Admin Remarks:</b> {escaped_remarks}" if escaped_remarks else ""
        
        # Notify employee
        notify_employee(
            telegram_id, 
            f"⚠️ <b>Fine Applied!</b>\n"
            f"An administrative fine has been added to your record.\n\n"
            f"• <b>Date & Time:</b> <code>{current_time_str}</code> (Record date: <code>{date}</code>)\n"
            f"• <b>Fine Amount:</b> <code>INR {fine_amount}</code>\n"
            f"• <b>Reason:</b> {escaped_reason}{remarks_block}",
            parse_mode="HTML"
        )
    else:
        db.delete_fine(telegram_id, date)
        # Format current timestamp for notification
        current_time_str = datetime.datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")
        # Notify employee
        notify_employee(
            telegram_id, 
            f"✅ <b>Fine Revoked!</b>\n"
            f"The administrative fine for <b>{date}</b> has been successfully removed.\n"
            f"• <b>Action Date & Time:</b> <code>{current_time_str}</code>",
            parse_mode="HTML"
        )
        
    return jsonify({"success": True})

@app.route('/api/attendance/edit', methods=['POST'])
def edit_attendance():
    """Manually add, update, or remove an employee's attendance session for a specific date."""
    data = request.get_json() or {}
    telegram_id = data.get("telegram_id")
    date_str = data.get("date")
    status = data.get("status") # "present" or "absent"

    if not telegram_id or not date_str or not status:
        return jsonify({"error": "Missing required fields: telegram_id, date, status"}), 400

    try:
        telegram_id = int(telegram_id)
    except ValueError:
        return jsonify({"error": "Invalid telegram_id"}), 400

    user = db.get_user(telegram_id)
    if not user:
        return jsonify({"error": f"Employee with Telegram ID {telegram_id} not found"}), 404

    conn = db.connect()
    cursor = conn.cursor()

    if status == "absent":
        try:
            # Delete any attendance, break, or in/out sessions for this user on this date
            cursor.execute("DELETE FROM attendance_sessions WHERE telegram_id = ? AND date = ?", (telegram_id, date_str))
            cursor.execute("DELETE FROM break_sessions WHERE telegram_id = ? AND date = ?", (telegram_id, date_str))
            cursor.execute("DELETE FROM in_out_sessions WHERE telegram_id = ? AND date = ?", (telegram_id, date_str))
            conn.commit()
            
            # Notify employee via bot
            notify_employee(
                telegram_id,
                f"ℹ️ *Attendance Update*\n"
                f"Your attendance for *{date_str}* has been set to *Absent* by the administrator.",
                parse_mode="Markdown"
            )
            return jsonify({"success": True, "message": "Attendance marked as absent (sessions deleted)"})
        except Exception as e:
            conn.rollback()
            return jsonify({"error": f"Database error: {str(e)}"}), 500

    # Otherwise status == "present"
    login_time = data.get("login_time")
    logout_time = data.get("logout_time")
    is_half_day = int(data.get("is_half_day", 0))
    fine_applied = int(data.get("fine_applied", 0))
    fine_amount = float(data.get("fine_amount", 0.0))
    fine_reason = data.get("fine_reason", "")
    remarks = data.get("remarks", "")

    if not login_time:
        return jsonify({"error": "Login time is required for present status"}), 400

    # Clean empty strings to None
    if not logout_time or logout_time.strip() == "":
        logout_time = None

    duration = 0
    if logout_time:
        duration = calculate_time_diff_seconds(login_time, logout_time)

    sess_status = "completed" if logout_time else "active"

    try:
        # Check if there are existing sessions
        cursor.execute(
            "SELECT id, fine_applied FROM attendance_sessions WHERE telegram_id = ? AND date = ? ORDER BY id ASC",
            (telegram_id, date_str)
        )
        rows = cursor.fetchall()
        
        had_fine_previously = False
        if rows:
            had_fine_previously = bool(rows[0]['fine_applied'])
            primary_session_id = rows[0]['id']
            # Delete any extra sessions if they exist
            if len(rows) > 1:
                other_ids = [r['id'] for r in rows[1:]]
                cursor.execute(
                    f"DELETE FROM attendance_sessions WHERE id IN ({','.join(['?']*len(other_ids))})",
                    other_ids
                )
            
            # Update primary session
            cursor.execute(
                """
                UPDATE attendance_sessions 
                SET login_time = ?, logout_time = ?, duration = ?, status = ?, 
                    is_half_day = ?, fine_applied = ?, fine_amount = ?, fine_reason = ?
                WHERE id = ?
                """,
                (login_time, logout_time, duration, sess_status,
                 is_half_day, fine_applied, fine_amount, fine_reason,
                 primary_session_id)
            )
        else:
            # Insert a brand new session
            cursor.execute(
                """
                INSERT INTO attendance_sessions 
                (telegram_id, username, name, date, login_time, logout_time, duration, status, is_half_day, fine_applied, fine_amount, fine_reason)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (telegram_id, user['username'] or "NoUsername", user['full_name'], date_str,
                 login_time, logout_time, duration, sess_status,
                 is_half_day, fine_applied, fine_amount, fine_reason)
            )
        conn.commit()

        # Synchronize audit fines table
        if fine_applied:
            db.create_fine(telegram_id, date_str, fine_amount, fine_reason)
        else:
            db.delete_fine(telegram_id, date_str)

        # Send Telegram notification
        status_label = "Half Day" if is_half_day else "Full Day"
        time_details = f"Login: `{login_time}`"
        if logout_time:
            time_details += f" | Logout: `{logout_time}`"
        
        if fine_applied:
            current_time_str = datetime.datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")
            import html
            escaped_reason = html.escape(fine_reason)
            escaped_remarks = html.escape(remarks) if remarks else ""
            remarks_block = f"\n💬 <b>Admin Remarks:</b> {escaped_remarks}" if escaped_remarks else ""
            
            notify_employee(
                telegram_id,
                f"⚠️ <b>Fine Notice (Applied/Reworked)</b>\n"
                f"An administrative fine has been applied or updated on your record.\n\n"
                f"• <b>Date & Time:</b> <code>{current_time_str}</code> (Record date: <code>{date_str}</code>)\n"
                f"• <b>Fine Amount:</b> <code>INR {fine_amount}</code>\n"
                f"• <b>Reason:</b> {escaped_reason}\n"
                f"• <b>Attendance Status:</b> {status_label} (Login: <code>{login_time}</code>" + 
                (f" | Logout: <code>{logout_time}</code>" if logout_time else "") + f"){remarks_block}",
                parse_mode="HTML"
            )
        elif had_fine_previously:
            current_time_str = datetime.datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")
            notify_employee(
                telegram_id,
                f"✅ <b>Fine Revoked!</b>\n"
                f"The administrative fine for <b>{date_str}</b> has been successfully removed.\n"
                f"• <b>Action Date & Time:</b> <code>{current_time_str}</code>",
                parse_mode="HTML"
            )
        else:
            # Just attendance adjusted without fine
            notify_employee(
                telegram_id,
                f"ℹ️ *Attendance Adjusted*\n"
                f"Your attendance for *{date_str}* has been updated by the administrator:\n"
                f"• Status: *Present* ({status_label})\n"
                f"• {time_details}",
                parse_mode="Markdown"
            )
        return jsonify({"success": True, "message": "Attendance updated successfully"})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database error: {str(e)}"}), 500

@app.route('/api/attendance/daily', methods=['GET'])
def get_daily_attendance():
    """Returns full attendance summary for ALL staff for a given date."""
    date_str = request.args.get('date', datetime.datetime.now(tz).strftime("%Y-%m-%d"))
    users = db.get_all_users()

    results = []
    present_count = 0
    absent_count = 0
    half_day_count = 0
    total_net_seconds = 0

    for u in users:
        telegram_id = u['telegram_id']
        summary = AttendanceReporter.get_employee_daily_summary(db, telegram_id, date_str)

        # Check if employee was present at all
        if summary['login_time'] == 'N/A':
            status = 'absent'
            absent_count += 1
        else:
            status = 'present'
            present_count += 1

        # Get is_half_day and fine from attendance session
        conn = db.connect()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, is_half_day, fine_applied, fine_amount, fine_reason FROM attendance_sessions WHERE telegram_id = ? AND date = ? ORDER BY id DESC LIMIT 1",
            (telegram_id, date_str)
        )
        sess_row = cursor.fetchone()
        is_half_day = 0
        fine_applied = 0
        fine_amount = 0.0
        fine_reason = ''
        session_id = None
        if sess_row:
            session_id, is_half_day, fine_applied, fine_amount, fine_reason = sess_row

        if is_half_day:
            half_day_count += 1

        total_net_seconds += summary['net_working_seconds']

        results.append({
            'telegram_id': telegram_id,
            'username': u['username'] or 'NoUsername',
            'full_name': u['full_name'],
            'role': u['role'],
            'shift_name': resolve_shift_name(u.get('shift_start') or DEFAULT_SHIFT_START, u.get('shift_end') or DEFAULT_SHIFT_END),
            'shift_start': u.get('shift_start') or DEFAULT_SHIFT_START,
            'shift_end': u.get('shift_end') or DEFAULT_SHIFT_END,
            'login_time': summary['login_time'],
            'logout_time': summary['logout_time'],
            'total_break_str': summary['total_break_str'],
            'net_working_str': summary['net_working_str'],
            'net_working_seconds': summary['net_working_seconds'],
            'status': status,
            'is_half_day': is_half_day,
            'session_id': session_id,
            'fine_applied': fine_applied,
            'fine_amount': fine_amount,
            'fine_reason': fine_reason or ''
        })

    # Sort: present first, then absent
    results.sort(key=lambda x: (0 if x['status'] == 'present' else 1, x['full_name']))

    avg_sec = total_net_seconds // present_count if present_count > 0 else 0

    return jsonify({
        'date': date_str,
        'records': results,
        'summary': {
            'total': len(results),
            'present': present_count,
            'absent': absent_count,
            'half_day': half_day_count,
            'avg_net_hours': format_seconds(avg_sec)
        }
    })

@app.route('/api/requests', methods=['GET'])
def get_requests():
    """Retrieve all submitted early logout requests."""
    rows = db.get_all_early_logout_requests()
    return jsonify(rows)

@app.route('/api/requests/<int:request_id>/review', methods=['POST'])
def review_request(request_id):
    """Process employee requests (approve/reject early checkout)."""
    data = request.get_json() or {}
    review_status = data.get("status")
    if review_status not in ["approved", "rejected"]:
        return jsonify({"error": "Invalid review status"}), 400
        
    req = db.get_early_logout_request(request_id)
    if not req:
        return jsonify({"error": "Request not found"}), 404
        
    telegram_id = req['telegram_id']
    date = req['date']
    reason = req['reason']
    
    # Update status
    success = db.update_early_logout_request_status(request_id, review_status)
    if not success:
        return jsonify({"error": "Database error"}), 500
        
    if review_status == "approved":
        # Mark attendance session as a half day
        db.set_attendance_half_day(telegram_id, date, 1)
        # Notify employee
        notify_employee(
            telegram_id,
            f"✅ *Early Logout Approved!*\n"
            f"Your early checkout request for *{date}* (Reason: {reason}) has been *APPROVED*.\n"
            f"The session is marked as a *Half Day*.",
            parse_mode="Markdown"
        )
    else:
        # Marked as normal
        db.set_attendance_half_day(telegram_id, date, 0)
        # Notify employee
        notify_employee(
            telegram_id,
            f"❌ *Early Logout Rejected!*\n"
            f"Your early checkout request for *{date}* (Reason: {reason}) has been *REJECTED*.\n"
            f"Please check with your administrator.",
            parse_mode="Markdown"
        )
        
    return jsonify({"success": True})

@app.route('/api/export/breaks', methods=['GET'])
def export_breaks_report():
    """Generates and exports an Excel break details report for a given date."""
    date_str = request.args.get('date', datetime.datetime.now(tz).strftime("%Y-%m-%d"))
    
    users = db.get_all_users()
    
    # We will gather break data for all employees
    records_data = []
    max_breaks = 3 # guarantee at least 3 break columns
    
    for u in users:
        # Skip admins unless they are professor_noxx
        if u['role'] == 'admin' and (not u['username'] or u['username'].lower() != 'professor_noxx'):
            continue
            
        telegram_id = u['telegram_id']
        name = u['full_name']
        
        # 1. Fetch break sessions
        brk_sessions = db.get_break_sessions_by_date(telegram_id, date_str)
        # 2. Fetch In/Out sessions
        move_sessions = db.get_in_out_sessions_by_date(telegram_id, date_str)
        
        # 3. Combine and sort chronologically
        all_breaks = []
        for b in brk_sessions:
            start_time = b['break_in_time']
            end_time = b['break_out_time'] or "Active"
            duration = b['duration'] or 0
            
            # If active and today, calculate dynamic duration
            if b['status'] == 'active' and date_str == datetime.datetime.now(tz).strftime("%Y-%m-%d"):
                from reports.reporter import calculate_time_diff_seconds, get_current_time_str
                duration = calculate_time_diff_seconds(start_time, get_current_time_str())
                
            all_breaks.append({
                'start': start_time,
                'end': end_time,
                'duration': duration,
                'type': 'Lunch'
            })
            
        for m in move_sessions:
            start_time = m['in_time']
            end_time = m['out_time'] or "Active"
            duration = m['duration'] or 0
            
            # If active and today, calculate dynamic duration
            if m['status'] == 'active' and date_str == datetime.datetime.now(tz).strftime("%Y-%m-%d"):
                from reports.reporter import calculate_time_diff_seconds, get_current_time_str
                duration = calculate_time_diff_seconds(start_time, get_current_time_str())
                
            all_breaks.append({
                'start': start_time,
                'end': end_time,
                'duration': duration,
                'type': 'Field'
            })
            
        # Sort chronologically by start time
        all_breaks.sort(key=lambda x: x['start'])
        
        # Calculate total break duration
        total_break_seconds = sum(b['duration'] for b in all_breaks)
        
        # Track max breaks to determine columns
        if len(all_breaks) > max_breaks:
            max_breaks = len(all_breaks)
            
        # Check if fine is applied
        conn = db.connect()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT fine_applied, fine_amount, fine_reason FROM attendance_sessions WHERE telegram_id = ? AND date = ? ORDER BY id DESC LIMIT 1",
            (telegram_id, date_str)
        )
        row = cursor.fetchone()
        fine_applied_str = "No"
        if row:
            fine_applied, fine_amount, fine_reason = row
            if fine_applied:
                fine_applied_str = f"Yes (INR {fine_amount:.0f})"
                if fine_reason:
                    fine_applied_str += f" - {fine_reason}"
                    
        records_data.append({
            'name': name,
            'telegram_id': telegram_id,
            'breaks': all_breaks,
            'total_break_seconds': total_break_seconds,
            'fine_applied_str': fine_applied_str
        })
        
    # Build headers dynamically
    headers = ["Staff Name", "Telegram ID"]
    for i in range(1, max_breaks + 1):
        headers.append(f"Break {i}")
    headers.extend(["Total Break Duration", "Fine Applied"])
    
    # Build rows
    rows = []
    for rd in records_data:
        row = [rd['name'], rd['telegram_id']]
        
        # Add breaks
        for i in range(max_breaks):
            if i < len(rd['breaks']):
                b = rd['breaks'][i]
                duration_min = b['duration'] // 60
                duration_sec = b['duration'] % 60
                
                if b['end'] == "Active":
                    time_str = f"{b['start']} - Active"
                else:
                    time_str = f"{b['start']} - {b['end']}"
                    
                # e.g., "12:30 - 13:15 (Lunch, 45m 0s)"
                row.append(f"{time_str} ({b['type']}, {duration_min}m {duration_sec}s)")
            else:
                row.append("—")
                
        # Total Break Duration
        from reports.reporter import format_seconds
        row.append(format_seconds(rd['total_break_seconds']))
        
        # Fine Applied
        row.append(rd['fine_applied_str'])
        
        rows.append(row)
        
    # Sort rows by Staff Name
    rows.sort(key=lambda x: x[0])
    
    # Create Excel workbook using openpyxl
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Break Details Report"
    
    # Title Row
    title_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"Staff Break Details Report  ·  Date: {date_str}  ·  Generated: {datetime.date.today()}")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Header Row
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='B0BEC5')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 30
    
    # Data Rows
    alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    data_align = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
                
    # Column Widths
    ws.column_dimensions['A'].width = 22 # Staff Name
    ws.column_dimensions['B'].width = 16 # Telegram ID
    for col_idx in range(3, 3 + max_breaks):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 30 # Break columns
        
    total_break_col_letter = get_column_letter(3 + max_breaks)
    fine_col_letter = get_column_letter(4 + max_breaks)
    
    ws.column_dimensions[total_break_col_letter].width = 20 # Total Break Duration
    ws.column_dimensions[fine_col_letter].width = 30 # Fine Applied
    
    # Freeze panes and Auto-filter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"staff_breaks_{date_str}.xlsx"
    )

@app.route('/api/export', methods=['GET'])
def export_records():
    """Export attendance data directly to CSV, Excel, or PDF."""
    export_format = request.args.get('format', 'csv').lower()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    telegram_id = request.args.get('telegram_id')
    shift_type_filter = request.args.get('shift_type', '').strip().lower()  # 'day', 'night', 'unassigned', or ''

    # Human-readable shift label used in filenames / title rows
    _shift_labels = {'day': 'Day Shift', 'night': 'Night Shift', 'unassigned': 'Unassigned'}
    shift_label = _shift_labels.get(shift_type_filter, 'All Shifts')

    # Query database records (include project and shift_type from users table)
    conn = db.connect()
    cursor = conn.cursor()
    
    query = """
        SELECT a.id, a.telegram_id, a.username, a.name, a.date, a.login_time, a.logout_time, 
               a.duration, a.is_half_day, a.fine_applied, a.fine_amount, a.fine_reason, 
               u.shift_start, u.shift_end, u.project, u.shift_type
        FROM attendance_sessions a
        JOIN users u ON a.telegram_id = u.telegram_id
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND a.date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND a.date <= ?"
        params.append(end_date)
    if telegram_id:
        query += " AND a.telegram_id = ?"
        params.append(int(telegram_id))

    # Shift-type filter: 'unassigned' means NULL or empty, otherwise match exactly
    if shift_type_filter == 'unassigned':
        query += " AND (u.shift_type IS NULL OR u.shift_type = '')"
    elif shift_type_filter in ('day', 'night'):
        query += " AND LOWER(u.shift_type) = ?"
        params.append(shift_type_filter)
        
    query += " ORDER BY a.date DESC, CAST(u.project AS INTEGER) ASC, u.project ASC, a.login_time DESC"
    cursor.execute(query, tuple(params))
    records = [dict(row) for row in cursor.fetchall()]
    
    # ── Headers in the exact requested order ───────────────────────────────
    headers = [
        "Date",
        "Staff Name",
        "Telegram ID",
        "Project Number",
        "Shift Type",
        "Login Time",
        "Logout Time",
        "Break Total Time",
        "Total Number of Breaks",
        "Fine Amount",
        "Attendance Status",
        "Shift Start Time",
        "Shift End Time",
        "Total Work Duration",
    ]
    
    # ── Build enriched rows with break stats ───────────────────────────────
    rows = []
    for r in records:
        tid = r['telegram_id']
        date_val = r['date']
        
        # Fetch all break and in/out sessions for this employee on this date
        brk_sessions = db.get_break_sessions_by_date(tid, date_val)
        move_sessions = db.get_in_out_sessions_by_date(tid, date_val)
        
        total_lunch_seconds = sum((b['duration'] or 0) if b['status'] == 'completed' else 0 for b in brk_sessions)
        total_move_seconds = sum((m['duration'] or 0) if m['status'] == 'completed' else 0 for m in move_sessions)
        total_break_seconds = total_lunch_seconds + total_move_seconds
        
        num_breaks = len(brk_sessions) + len(move_sessions)
        break_total_str = format_seconds(total_break_seconds)
        
        # Total work duration: raw session duration minus total break time
        raw_duration = r['duration'] or 0
        net_work_seconds = max(raw_duration - total_break_seconds, 0)
        work_duration_str = format_seconds(net_work_seconds) if raw_duration else "Active"
        
        # Attendance status label
        attendance_status = "Half Day" if r['is_half_day'] else "Full Day"
        
        # Fine amount (0.0 if no fine applied)
        fine_amount = r['fine_amount'] if r['fine_applied'] else 0.0
        
        # Friendly shift type label
        raw_shift = (r.get('shift_type') or '').strip().lower()
        shift_type_display = 'Day Shift' if raw_shift == 'day' else 'Night Shift' if raw_shift == 'night' else 'Unassigned'

        rows.append([
            date_val,                           # Date
            r['name'],                          # Staff Name
            r['telegram_id'],                   # Telegram ID
            r['project'] or "N/A",             # Project Number
            shift_type_display,                 # Shift Type
            r['login_time'],                    # Login Time
            r['logout_time'] or "N/A",          # Logout Time
            break_total_str,                    # Break Total Time
            num_breaks,                         # Total Number of Breaks
            fine_amount,                        # Fine Amount
            attendance_status,                  # Attendance Status
            r['shift_start'] or "N/A",         # Shift Start Time
            r['shift_end'] or "N/A",           # Shift End Time
            work_duration_str,                  # Total Work Duration
        ])
        
    # ─── Generate CSV ────────────────────────────────────────────────────────
    if export_format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        shift_slug = shift_type_filter if shift_type_filter else 'all'
        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"attendance_{shift_slug}_shift_{datetime.date.today()}.csv"
        )
        
    # ─── Generate Excel (using openpyxl) ────────────────────────────────────
    elif export_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staff Attendance Report"
            
            # ── Title row ──
            title_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1,
                                 value=f"Staff Attendance Report  ·  {shift_label}  ·  Generated: {datetime.date.today()}")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22
            
            # ── Header row ──
            header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin = Side(style='thin', color='B0BEC5')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[2].height = 32
            
            # ── Data rows ──
            alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
            data_align = Alignment(horizontal='center', vertical='center')
            
            for row_idx, row_data in enumerate(rows, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = data_align
                    cell.border = border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
            
            # ── Column widths (characters) ── [14 cols]
            col_widths = [12, 22, 16, 14, 14, 12, 12, 16, 20, 12, 18, 14, 14, 18]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
                
            # ── Freeze panes below header ──
            ws.freeze_panes = "A3"
            
            # ── Auto-filter on header row ──
            ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
            
            excel_stream = io.BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            
            shift_slug = shift_type_filter if shift_type_filter else 'all'
            return send_file(
                excel_stream,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"attendance_{shift_slug}_shift_{datetime.date.today()}.xlsx"
            )
        except ImportError:
            return jsonify({"error": "openpyxl not installed on host"}), 500
            
    # ─── Generate PDF (using reportlab) ─────────────────────────────────────
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A3, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            pdf_stream = io.BytesIO()
            doc = SimpleDocTemplate(
                pdf_stream,
                pagesize=landscape(A3),
                rightMargin=15, leftMargin=15, topMargin=18, bottomMargin=15
            )
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1F3864'),
                spaceAfter=4
            )
            sub_style = ParagraphStyle(
                'SubStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#555555'),
                spaceAfter=10
            )
            
            story = []
            story.append(Paragraph("Staff Attendance Report", title_style))
            story.append(Paragraph(
                f"Shift: {shift_label}  \u00b7  Generated: {datetime.date.today()}  \u00b7  Total Records: {len(rows)}",
                sub_style
            ))
            story.append(Spacer(1, 4))
            
            # All 14 columns
            pdf_table_rows = [headers]
            for row_data in rows:
                pdf_table_rows.append([
                    str(row_data[0]),         # Date
                    str(row_data[1])[:22],    # Staff Name
                    str(row_data[2]),         # Telegram ID
                    str(row_data[3])[:14],    # Project Number
                    str(row_data[4]),         # Shift Type
                    str(row_data[5]),         # Login Time
                    str(row_data[6]),         # Logout Time
                    str(row_data[7]),         # Break Total Time
                    str(row_data[8]),         # Total Number of Breaks
                    str(row_data[9]),         # Fine Amount
                    str(row_data[10]),        # Attendance Status
                    str(row_data[11]),        # Shift Start Time
                    str(row_data[12]),        # Shift End Time
                    str(row_data[13]),        # Total Work Duration
                ])
            
            # Column widths (points) – tuned to fit A3 landscape (~1155 pt usable) [14 cols]
            col_widths_pt = [50, 82, 66, 54, 52, 46, 46, 56, 50, 46, 58, 52, 52, 58]
            
            table = Table(pdf_table_rows, colWidths=col_widths_pt, repeatRows=1)
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND',     (0, 0), (-1, 0), colors.HexColor('#1F3864')),
                ('TEXTCOLOR',      (0, 0), (-1, 0), colors.white),
                ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',       (0, 0), (-1, 0), 7.5),
                ('TOPPADDING',     (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING',  (0, 0), (-1, 0), 7),
                ('ALIGN',          (0, 0), (-1, 0), 'CENTER'),
                ('LINEBELOW',      (0, 0), (-1, 0), 1.2, colors.HexColor('#1F3864')),
                # Data row styling
                ('FONTNAME',       (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE',       (0, 1), (-1, -1), 6.5),
                ('ALIGN',          (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',     (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING',  (0, 1), (-1, -1), 4),
                # Alternating row colours
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
                # Grid
                ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#B0BEC5')),
            ]))
            
            story.append(table)
            doc.build(story)
            pdf_stream.seek(0)
            
            shift_slug = shift_type_filter if shift_type_filter else 'all'
            return send_file(
                pdf_stream,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"attendance_{shift_slug}_shift_{datetime.date.today()}.pdf"
            )
        except ImportError:
            return jsonify({"error": "reportlab not installed on host"}), 500
            
    return jsonify({"error": "Unsupported format"}), 400


# ──────────────────────────────────────────────────────────────
# Permission Request API Routes
# ──────────────────────────────────────────────────────────────

@app.route('/api/permissions', methods=['GET'])
def get_permissions():
    """
    Returns all permission requests.
    Optional query params:
      ?status=pending|approved|rejected
      ?date=YYYY-MM-DD  (filters by the permission date)
    """
    status_filter = request.args.get('status', None)
    date_filter = request.args.get('date', None)

    try:
        records = db.get_all_permission_requests(status=status_filter)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Optional date filter
    if date_filter:
        records = [r for r in records if r.get('date') == date_filter]

    return jsonify(records)


@app.route('/api/permissions/<int:request_id>/approve', methods=['POST'])
def approve_permission(request_id: int):
    """Approve a permission request from the dashboard."""
    data = request.get_json(silent=True) or {}
    approver_name = data.get('approver_name', 'Admin (Dashboard)')

    req = db.get_permission_request(request_id)
    if not req:
        return jsonify({"error": "Permission request not found"}), 404

    if req['status'] != 'pending':
        return jsonify({"error": f"Request is already {req['status']}"}), 409

    ok = db.update_permission_request_status(
        request_id=request_id,
        status='approved',
        approver_id=None,
        approver_name=approver_name,
    )

    if ok:
        # Unconditionally reverse fine and upgrade to Full Day on permission approval
        try:
            telegram_id = req['telegram_id']
            date = req['date']

            # Remove any fine on record for this date
            db.delete_fine(telegram_id, date)

            # Flip all sessions on that date to Full Day and clear fine fields
            conn = db.connect()
            conn.execute(
                """
                UPDATE attendance_sessions 
                SET is_half_day = 0, fine_applied = 0, fine_amount = 0.0, fine_reason = ''
                WHERE telegram_id = ? AND date = ?
                """,
                (telegram_id, date)
            )
            conn.commit()
        except Exception as recalc_err:
            print(f"⚠️ Fine reversal error: {recalc_err}")

        # Notify employee via Telegram
        send_and_record_permission_notification(req, 'approved', approver_name)

    return jsonify({"success": ok, "status": "approved"})


@app.route('/api/permissions/<int:request_id>/reject', methods=['POST'])
def reject_permission(request_id: int):
    """Reject a permission request from the dashboard."""
    data = request.get_json(silent=True) or {}
    approver_name = data.get('approver_name', 'Admin (Dashboard)')

    req = db.get_permission_request(request_id)
    if not req:
        return jsonify({"error": "Permission request not found"}), 404

    if req['status'] != 'pending':
        return jsonify({"error": f"Request is already {req['status']}"}), 409

    ok = db.update_permission_request_status(
        request_id=request_id,
        status='rejected',
        approver_id=None,
        approver_name=approver_name,
    )

    if ok:
        # Notify employee via Telegram
        send_and_record_permission_notification(req, 'rejected', approver_name)

    return jsonify({"success": ok, "status": "rejected"})


@app.route('/api/permissions/<int:request_id>/retry-notification', methods=['POST'])
def retry_permission_notification(request_id: int):
    """Retry sending the Telegram notification for a permission decision."""
    req = db.get_permission_request(request_id)
    if not req:
        return jsonify({"error": "Permission request not found"}), 404
        
    if req['status'] == 'pending':
        return jsonify({"error": "Cannot notify for a pending request"}), 400

    approver_name = req.get('approver_name') or 'Admin (Dashboard)'
    success = send_and_record_permission_notification(req, req['status'], approver_name)
    
    return jsonify({"success": success})


if __name__ == '__main__':
    # Launch backend API server
    # NOTE: this dev-server entrypoint is for local development only. In
    # production (Render, etc.) run via a WSGI server instead, e.g.:
    #   gunicorn dashboard.app:app --bind 0.0.0.0:$PORT
    port = int(os.getenv("PORT", 5000))
    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    print("---------------------------------------------")
    print("🖥️ Starting Attendance Web Dashboard Server")
    print(f"🔗 Admin Link: http://localhost:{port}")
    print("---------------------------------------------")
    app.run(host='0.0.0.0', port=port, debug=debug_mode)
